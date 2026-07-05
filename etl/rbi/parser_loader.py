import os
import openpyxl
from datetime import datetime
import psycopg2
import json


def parse_and_load():
    # Load the workbook
    # filename = 'rbi-data-sample.xlsx'
    filename = 'all-data.xlsx'
    
    print("loading file: ", filename)
    wb = openpyxl.load_workbook(filename, data_only=True)
    print("file successfully loaded")
    
    

    data_list = []  # Re-initialize to clear any previously added data

    # Process each sheet again with additional corrections
    print("loaded sheet with names: ", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        month, year = sheet_name.rsplit(' ', 1)
        for row in sheet.iter_rows(min_row=7, values_only=True):
            if row[0] is None:
                continue

            if isinstance(row[0], datetime):
                day = row[0].day
            else:
                try:
                    parsed_date = datetime.strptime(row[0], "%d-%m-%Y")
                    day = parsed_date.day
                except ValueError:
                    continue

            for col in range(2, len(row), 2):
                product = sheet.cell(row=5, column=col).value
                volume = row[col - 1] if col - 1 < len(row) else None
                value = row[col] if col < len(row) else None
                if product and volume and value:
                    data_list.append({
                        "Product": product,
                        "Volume": str(volume),
                        "Value": str(value),
                        "Day": str(day),
                        "Month": month,
                        "Year": year,
                    })
                    
    print("Parsing Complete. Rows processed = ", len(data_list))

    conn = None
    try:
        conn = psycopg2.connect(
            host="localhost",
            user="admin",
            password=os.environ["DB_PASSWORD"],
            dbname="npci",
            port=5432
        )
        for item in data_list:
            try:
                product = item['Product']
                category = item['Category']
                sub_category = item['Sub-Category']
                try:
                    if item['Volume'] == 'h':
                        volume = 0
                    else:
                        volume = float(item['Volume'])
                    
                except ValueError:
                    print(f"Error occured due to invalid number format for volume: {item}. Using 0")
                    volume = 0 
                try:  
                    if item['Value'] == 'h':
                        value = 0
                    else:
                        value = float(item['Volume'])
                except ValueError:
                    print(f"Error occured due to invalid number format for value: {item}. Using 0")
                    value = 0 
                date = datetime(int(item['Year']), datetime.strptime(item['Month'], "%B").month, int(item['Day']))

                with conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "INSERT INTO rbi_dev.daily_product_statistics (product, category, sub_category,,volume, value, date) VALUES (%s, %s, %s, %s)",
                            (product, category, sub_category, volume, value, date)
                        )
            except Exception as insert_error:
                print(f"Warning: Failed to insert record {product}, {volume}, {value}, {date}. Error: {insert_error}")
    finally:
        if conn is not None:
            conn.close()

    print(f"Successfully processed {len(data_list)} records.")




if __name__ == "__main__":
    print("Started Parsing and Loading")
    parse_and_load()
    print("Data loading complete.")