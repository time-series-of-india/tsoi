import openpyxl
from datetime import datetime
import json
from collections import defaultdict

CARD_BASES = ['Credit Card', 'Debit Card', 'Prepaid Payment Instruments (PPIs) Card']

def _is_card_split(sheet):
    """Return True if this sheet uses separate at-PoS / at-e-Commerce columns for cards."""
    return any(
        sheet.cell(row=6, column=col).value == 'at PoS'
        for col in range(1, min(sheet.max_column + 1, 60))
    )

def parse_data(file_path, product_dict_path):
    with open(product_dict_path, 'r') as file:
        product_dictionary = json.load(file)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    data_list = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        month, year = sheet_name.strip().rsplit(' ', 1)
        is_separated = _is_card_split(sheet)

        for row in sheet.iter_rows(min_row=7, values_only=True):
            if row[0] is None:
                continue

            if isinstance(row[0], datetime):
                day = row[0].day
            else:
                parsed_date = None
                for fmt in ("%d-%m-%Y", "%B %d, %Y"):
                    try:
                        parsed_date = datetime.strptime(row[0], fmt)
                        break
                    except (ValueError, TypeError):
                        pass
                if parsed_date is None:
                    continue
                day = parsed_date.day

            for col in range(2, len(row), 2):
                product = sheet.cell(row=5, column=col).value

                if is_separated:
                    # e-Commerce column: row5 is empty, inherit card name from 2 cols left
                    if product is None and col > 2 and sheet.cell(row=5, column=col - 2).value in CARD_BASES:
                        product = sheet.cell(row=5, column=col - 2).value

                # Normalize all card variants to base product name
                if product:
                    for base in CARD_BASES:
                        if product.startswith(base):
                            product = base
                            break

                volume = row[col - 1] if col - 1 < len(row) else None
                value = row[col] if col < len(row) else None

                if product and volume and value:
                    if product in product_dictionary:
                        category = product_dictionary[product].get("category", "Unknown")
                        sub_category = product_dictionary[product].get("sub-category", "Unknown")
                    else:
                        category = "Unknown"
                        sub_category = "Unknown"
                        print(f"Product '{product}' not found in product dictionary")

                    data_list.append({
                        "Product": product,
                        "Category": category,
                        "Sub-Category": sub_category,
                        "Volume": str(volume),
                        "Value": str(value),
                        "Day": str(day),
                        "Month": month,
                        "Year": year,
                    })

    return data_list


def process_data(data_list):
    """Sum PoS + e-Commerce into a single row per card product per day."""
    CARD_BASES_SET = set(CARD_BASES)
    combined = defaultdict(lambda: {"Volume": 0.0, "Value": 0.0, "meta": None})
    result = []

    for item in data_list:
        if item["Product"] in CARD_BASES_SET:
            key = (item["Product"], item["Day"], item["Month"], item["Year"])
            vol = float(item["Volume"]) if item["Volume"] != 'h' else 0.0
            val = float(item["Value"]) if item["Value"] != 'h' else 0.0
            combined[key]["Volume"] += vol
            combined[key]["Value"] += val
            if combined[key]["meta"] is None:
                combined[key]["meta"] = item
        else:
            result.append(item)

    for (product, day, month, year), data in combined.items():
        meta = data["meta"]
        result.append({
            "Product": product,
            "Category": meta["Category"],
            "Sub-Category": meta["Sub-Category"],
            "Volume": str(data["Volume"]),
            "Value": str(data["Value"]),
            "Day": day,
            "Month": month,
            "Year": year,
        })

    return result


def write_to_file(data_list, file_path):
    with open(file_path, 'w') as file:
        json.dump(data_list, file)
